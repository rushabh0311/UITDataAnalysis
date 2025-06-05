import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Input file paths
file_path = input("Enter the excel file name (with .xlsx extension): ")
sheet_name = input("Enter the sheet name: ")
output_file = input("Enter the output Excel file name (with .xlsx extension): ")

print("Please wait, your data is processing....")

# Load the data from Excel
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Ensure numeric conversion and handle errors
df.iloc[:, 1] = pd.to_numeric(df.iloc[:, 1], errors='coerce')
df['Motor_Current_LH'] = pd.to_numeric(df['Motor_Current_LH'], errors='coerce')
df['supply voltage'] = pd.to_numeric(df['supply voltage'], errors='coerce')
df['Angle'] = pd.to_numeric(df['Angle'], errors='coerce')

# Columns assignment
time_col = pd.to_numeric(df.iloc[:, 0], errors='coerce')
voltage_col = df.iloc[:, 1]
current_col = df['Motor_Current_LH']
supply_voltage_col = df['supply voltage']
angle_col = df['Angle']

results = []
selected_data = []
cycles = []
start_index = 0

# Detecting and processing cycles
while start_index < len(voltage_col):
    # Detect Open Cycle Start (voltage rising from 0 to >=1)
    while start_index < len(voltage_col) and voltage_col.iloc[start_index] < 1:
        start_index += 1
    if start_index >= len(voltage_col):
        break

    starting_point = time_col.iloc[start_index]

    closing_end_index = start_index
    while closing_end_index < len(voltage_col) and voltage_col.iloc[closing_end_index] >= 1:
        closing_end_index += 1
    if closing_end_index >= len(voltage_col):
        break

    closing_end_time = time_col.iloc[closing_end_index]
    time_lapse = closing_end_time - starting_point

    if time_lapse < 2:
        start_index = closing_end_index + 1
        continue

    cycles.append((starting_point, closing_end_time))

    cycle_voltage = voltage_col.iloc[start_index:closing_end_index].abs()
    cycle_current = current_col.iloc[start_index:closing_end_index].abs()
    avg_supply_voltage = supply_voltage_col.iloc[start_index:closing_end_index].mean()
    angle_value = angle_col.iloc[closing_end_index + 2] if closing_end_index + 2 < len(angle_col) else None

    results.append({
        "Cycle Type": "Open",
        "Starting Point(sec)": starting_point,
        "Highest Voltage(V)": cycle_voltage.max(),
        "Peak Current(Amp)": cycle_current.max(),
        "Running Current(Amp)": cycle_current.mean(),
        "Avg Supply Voltage(V)": avg_supply_voltage,
        "End Time(sec)": closing_end_time,
        "Time Lapse(sec)": time_lapse,
        "Angle(deg)": angle_value
    })

    selected_data.append(df.iloc[start_index:closing_end_index])
    start_index = closing_end_index + 1
 # Detect Close Cycle Start (voltage dropping below -1)
    while start_index < len(voltage_col) and voltage_col.iloc[start_index] >= -1:
        start_index += 1
    if start_index >= len(voltage_col):  # Exit if we reach the end
        break

    starting_point = time_col.iloc[start_index]

    closing_end_index = start_index
    while closing_end_index < len(voltage_col) and voltage_col.iloc[closing_end_index] < -1:
        closing_end_index += 1
    if closing_end_index >= len(voltage_col):
        break

    closing_end_time = time_col.iloc[closing_end_index]
    cycles.append((starting_point, closing_end_time))
    time_lapse = closing_end_time - starting_point

    # Absolute values during the cycle range for the Close Cycle
    cycle_voltage = voltage_col.iloc[start_index:closing_end_index].abs()
    cycle_current = current_col.iloc[start_index:closing_end_index].abs()
    peak_current = cycle_current.max()
    avg_current = cycle_current.mean()

    cycle_supply_voltage = supply_voltage_col.iloc[start_index:closing_end_index].abs()
    avg_supply_voltage = cycle_supply_voltage.mean()

    angle_point_index = closing_end_index + 2
    angle_value = angle_col.iloc[angle_point_index] if angle_point_index < len(angle_col) else None

    results.append({
        "Cycle Type": "Close",
        "Starting Point(sec)": starting_point,
        "Highest Voltage(V)": cycle_voltage.max(),
        "Peak Current(Amp)": peak_current,
        "Running Current(Amp)": avg_current,
        "Avg Supply Voltage(V)": avg_supply_voltage,
        "End Time(sec)": closing_end_time,
        "Time Lapse(sec)": time_lapse,
        "Angle(deg)": angle_value
    })

    selected_data.append(df.iloc[start_index:closing_end_index])
    start_index = closing_end_index  # Move to the next point after the current cycle

result_df = pd.DataFrame(results)
selected_data_df = pd.concat(selected_data)

print(result_df)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='Results', index=False)
    selected_data_df.to_excel(writer, sheet_name='Selected Data', index=False)


# Generate and save graphs
for i in range(0, len(cycles), 2):
    if i + 1 >= len(cycles):
        break

    start_open = cycles[i][0]
    end_close = cycles[i + 1][1]

    mask = (time_col >= start_open) & (time_col <= end_close)
    filtered_time = time_col[mask]
    filtered_angle = angle_col[mask]
    filtered_voltage = abs(voltage_col[mask])
    filtered_current = abs(current_col[mask])
    avg_supply_voltage = supply_voltage_col[mask].mean()

     # Ensure graphs start and end at zero
    filtered_time = [filtered_time.min()] + list(filtered_time) + [filtered_time.max()]
    filtered_angle = [0] + list(filtered_angle) + [0]
    filtered_voltage = [0] + list(filtered_voltage) + [0]
    filtered_current = [0] + list(filtered_current) + [0]

    plt.figure(figsize=(15, 10))
    plt.suptitle(f'Cycle {i // 2 + 1}Supply_Voltage{avg_supply_voltage:.2f}', fontsize=16)

    plt.subplot(3, 1, 1)
    plt.plot(filtered_time, filtered_angle, label='Angle', color='blue')
    plt.xlabel('Time(sec)')
    plt.ylabel('Angle(deg)')
    plt.grid(True)

    plt.subplot(3, 1, 2)
    plt.plot(filtered_time, filtered_voltage, label='Voltage', color='red')
    plt.xlabel('Time(sec)')
    plt.ylabel('Voltage(V)')
    plt.grid(True)

    plt.subplot(3, 1, 3)
    plt.plot(filtered_time, filtered_current, label='Current', color='green')
    plt.xlabel('Time(sec)')
    plt.ylabel('Current(Amp)')
    plt.grid(True)

    plt.tight_layout(rect=[0, 0, 1, 0.96])

    graph_path = f'cycle_{i // 2 + 1}supply_voltage{avg_supply_voltage:.2f}.png'
    plt.savefig(graph_path)
    plt.close()

    book = load_workbook(output_file)
    sheet_name = f'Supply_{avg_supply_voltage:.2f}'
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet = book[sheet_name]

    img = Image(graph_path)
    img.anchor = f'A{i // 2 + 2}'
    sheet.add_image(img)
    book.save(output_file)

print("Processing complete. Results saved to:", output_file)
